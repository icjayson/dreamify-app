from __future__ import annotations

import hashlib
import json
import os
import zipfile
from pathlib import Path

import pytest
import xlwt
from conftest import profile_request
from openpyxl import Workbook

from runner.errors import RunnerError
from runner.loaders import load_asset, verify_excel_archive
from runner.profiling import build_profile
from runner.schemas import AssetRequest, ProfileRequest


def test_profiles_utf8_csv_with_strict_caps(csv_workspace: tuple[Path, dict]) -> None:
    root, asset = csv_workspace
    result = build_profile(profile_request(asset), root)

    assert result["schema_version"] == "1"
    dataset = result["datasets"][0]
    assert dataset["row_count"] == 3
    assert dataset["column_count"] == 4
    assert {column["name"] for column in dataset["columns"]} == {
        "date",
        "region",
        "revenue",
        "orders",
    }


def test_rejects_hash_mismatch(csv_workspace: tuple[Path, dict]) -> None:
    root, raw_asset = csv_workspace
    raw_asset["sha256"] = "0" * 64
    asset = AssetRequest.from_mapping(raw_asset)
    with pytest.raises(RunnerError) as captured:
        load_asset(root, asset)
    assert captured.value.code == "ASSET_HASH_MISMATCH"


def test_rejects_workspace_traversal(csv_workspace: tuple[Path, dict]) -> None:
    _, raw_asset = csv_workspace
    raw_asset["relative_path"] = "../sales.csv"
    with pytest.raises(RunnerError) as captured:
        AssetRequest.from_mapping(raw_asset)
    assert captured.value.code == "UNSAFE_PATH"


def test_rejects_symlinked_asset(
    csv_workspace: tuple[Path, dict], tmp_path: Path
) -> None:
    root, raw_asset = csv_workspace
    external = tmp_path.parent / f"external-{tmp_path.name}.csv"
    external.write_text("a\n1\n", encoding="utf-8")
    path = root / "input" / "sales.csv"
    path.unlink()
    try:
        os.symlink(external, path)
        content = external.read_bytes()
        raw_asset["size_bytes"] = len(content)
        raw_asset["sha256"] = hashlib.sha256(content).hexdigest()
        with pytest.raises(RunnerError) as captured:
            load_asset(root, AssetRequest.from_mapping(raw_asset))
        assert captured.value.code == "UNSAFE_PATH"
    finally:
        external.unlink(missing_ok=True)


def test_rejects_nested_json(tmp_path: Path) -> None:
    input_directory = tmp_path / "input"
    input_directory.mkdir()
    content = json.dumps([{"name": "A", "nested": {"value": 1}}]).encode()
    path = input_directory / "nested.json"
    path.write_bytes(content)
    asset = AssetRequest.from_mapping(
        {
            "asset_id": "asset-json",
            "object_id": "object-json",
            "file_name": "nested.json",
            "format": "json",
            "media_type": "application/json",
            "size_bytes": len(content),
            "sha256": hashlib.sha256(content).hexdigest(),
            "relative_path": "input/nested.json",
        }
    )
    with pytest.raises(RunnerError) as captured:
        load_asset(tmp_path, asset)
    assert captured.value.code == "NESTED_JSON_UNSUPPORTED"


def test_profiles_flat_json_array(tmp_path: Path) -> None:
    input_directory = tmp_path / "input"
    input_directory.mkdir()
    content = json.dumps(
        [
            {"region": "North", "revenue": 1000},
            {"region": "South", "revenue": 1200},
        ]
    ).encode()
    path = input_directory / "sales.json"
    path.write_bytes(content)
    request = ProfileRequest.from_mapping(
        {
            "schema_version": "1",
            "run_id": "run-json",
            "assets": [
                {
                    "asset_id": "asset-json",
                    "object_id": "object-json",
                    "file_name": "sales.json",
                    "format": "json",
                    "media_type": "application/json",
                    "size_bytes": len(content),
                    "sha256": hashlib.sha256(content).hexdigest(),
                    "relative_path": "input/sales.json",
                }
            ],
        }
    )
    result = build_profile(request, tmp_path)
    assert result["datasets"][0]["row_count"] == 2
    assert result["datasets"][0]["column_count"] == 2


def test_rejects_more_than_two_hundred_columns(tmp_path: Path) -> None:
    input_directory = tmp_path / "input"
    input_directory.mkdir()
    content = (
        ",".join(f"column_{index}" for index in range(201))
        + "\n"
        + ",".join("1" for _ in range(201))
        + "\n"
    ).encode()
    path = input_directory / "wide.csv"
    path.write_bytes(content)
    asset = AssetRequest.from_mapping(
        {
            "asset_id": "asset-wide",
            "object_id": "object-wide",
            "file_name": "wide.csv",
            "format": "csv",
            "media_type": "text/csv",
            "size_bytes": len(content),
            "sha256": hashlib.sha256(content).hexdigest(),
            "relative_path": "input/wide.csv",
        }
    )
    with pytest.raises(RunnerError) as captured:
        load_asset(tmp_path, asset)
    assert captured.value.code == "COLUMN_LIMIT_EXCEEDED"


def test_selects_first_nonempty_visible_xlsx_sheet(tmp_path: Path) -> None:
    input_directory = tmp_path / "input"
    input_directory.mkdir()
    path = input_directory / "book.xlsx"
    workbook = Workbook()
    hidden = workbook.active
    hidden.title = "Hidden"
    hidden.sheet_state = "hidden"
    visible = workbook.create_sheet("Sales")
    visible.append(["region", "revenue"])
    visible.append(["North", 1000])
    workbook.save(path)
    content = path.read_bytes()
    request = ProfileRequest.from_mapping(
        {
            "schema_version": "1",
            "run_id": "run-xlsx",
            "assets": [
                {
                    "asset_id": "asset-xlsx",
                    "object_id": "object-xlsx",
                    "file_name": "book.xlsx",
                    "format": "xlsx",
                    "media_type": (
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                    "size_bytes": len(content),
                    "sha256": hashlib.sha256(content).hexdigest(),
                    "relative_path": "input/book.xlsx",
                }
            ],
        }
    )
    result = build_profile(request, tmp_path)
    assert result["datasets"][0]["sheet_name"] == "Sales"


def test_profiles_legacy_xls_with_pinned_xlrd(tmp_path: Path) -> None:
    input_directory = tmp_path / "input"
    input_directory.mkdir()
    path = input_directory / "legacy.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sales")
    for column, value in enumerate(("region", "revenue")):
        sheet.write(0, column, value)
    sheet.write(1, 0, "North")
    sheet.write(1, 1, 1000)
    workbook.save(str(path))
    content = path.read_bytes()
    request = ProfileRequest.from_mapping(
        {
            "schema_version": "1",
            "run_id": "run-xls",
            "assets": [
                {
                    "asset_id": "asset-xls",
                    "object_id": "object-xls",
                    "file_name": "legacy.xls",
                    "format": "xls",
                    "media_type": "application/vnd.ms-excel",
                    "size_bytes": len(content),
                    "sha256": hashlib.sha256(content).hexdigest(),
                    "relative_path": "input/legacy.xls",
                }
            ],
        }
    )
    result = build_profile(request, tmp_path)
    assert result["datasets"][0]["sheet_name"] == "Sales"
    assert result["datasets"][0]["row_count"] == 1


def test_rejects_unsafe_xlsx_compression_ratio(tmp_path: Path) -> None:
    path = tmp_path / "bomb.xlsx"
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"0" * (2 * 1024 * 1024))
    with pytest.raises(RunnerError) as captured:
        verify_excel_archive(path)
    assert captured.value.code == "UNSAFE_ARCHIVE"
